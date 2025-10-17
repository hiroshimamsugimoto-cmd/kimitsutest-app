import streamlit as st
import openpyxl
from datetime import datetime
from io import BytesIO
import base64
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication

# === 設定 ===
TEMPLATE = "気密試験記録.xlsx"

st.title("📘 気密試験記録 入力フォーム")

# --- メール設定の保存用 ---
if "email" not in st.session_state:
    st.session_state["email"] = None
if "app_password" not in st.session_state:
    st.session_state["app_password"] = None

# --- Outlook送信設定（初回のみ） ---
with st.expander("📧 Outlook送信設定（初回のみ入力）", expanded=True):
    email = st.text_input("自分のOutlookメールアドレス", value=st.session_state["email"] or "")
    app_password = st.text_input("アプリパスワード（Outlook）", type="password", value=st.session_state["app_password"] or "")
    save_setting = st.checkbox("メール設定を保存する（ブラウザを閉じても保持されます）")

    if st.button("保存"):
        st.session_state["email"] = email
        st.session_state["app_password"] = app_password
        st.success("✅ メール設定を保存しました。")
        st.rerun()

    if st.button("設定をリセット"):
        st.session_state["email"] = None
        st.session_state["app_password"] = None
        st.warning("⚠ メール設定を削除しました。再入力が必要です。")
        st.rerun()

# --- 入力項目 ---
st.header("入力項目")
系統名 = st.text_input("系統名")
試験圧力 = st.text_input("試験圧力 (MPa)")
試験範囲 = st.text_input("試験範囲")
試験媒体 = st.text_input("試験媒体")
放置時間 = st.text_input("放置時間 (h)", placeholder="例：10min以上")
使用機器No = st.text_input("使用圧力計機器No.")
測定場所 = st.text_input("測定場所")

# --- 開始日時 ---
st.subheader("開始日時")
col1, col2, col3 = st.columns([2, 1, 1])
with col1:
    開始日 = st.date_input("日付", key="start_date")
with col2:
    開始時 = st.text_input("時", value="", key="start_hour")
with col3:
    開始分 = st.text_input("分", value="", key="start_minute")

# --- 終了日時 ---
st.subheader("終了日時")
col4, col5, col6 = st.columns([2, 1, 1])
with col4:
    終了日 = st.date_input("日付", key="end_date")
with col5:
    終了時 = st.text_input("時", value="", key="end_hour")
with col6:
    終了分 = st.text_input("分", value="", key="end_minute")

# --- 測定値入力 ---
st.subheader("測定値入力")
col5, col6 = st.columns(2)
with col5:
    P1 = st.text_input("開始圧力 (MPa)", placeholder="例：0.8760")
with col6:
    T1 = st.text_input("開始温度 (℃)", placeholder="例：20.1")

col7, col8 = st.columns(2)
with col7:
    P2p = st.text_input("終了圧力 (MPa)", placeholder="例：0.8756")
with col8:
    T2 = st.text_input("終了温度 (℃)", placeholder="例：19.3")

試験実施者 = st.text_input("試験実施者")

# --- 数値変換 ---
def safe_float(v):
    try:
        return float(v.strip()) if v else None
    except:
        return None

P1 = safe_float(P1)
T1 = safe_float(T1)
P2p = safe_float(P2p)
T2 = safe_float(T2)

# --- 判定・保存 ---
if st.button("判定・保存"):
    if None in (P1, T1, P2p, T2):
        st.warning("⚠ 圧力・温度のすべてを入力してください。")
    else:
        try:
            開始日時 = datetime.combine(開始日, datetime.strptime(f"{開始時:02d}:{開始分:02d}", "%H:%M").time())
            終了日時 = datetime.combine(終了日, datetime.strptime(f"{終了時:02d}:{終了分:02d}", "%H:%M").time())

            T1_K = T1 + 273.15
            T2_K = T2 + 273.15
            P2_corr = P2p * (T1_K / T2_K)
            ΔP = P2_corr - P1
            判定範囲 = P1 * 0.01
            合否 = "合格" if abs(ΔP) <= 判定範囲 else "不合格"
            色 = "green" if 合否 == "合格" else "red"

            st.markdown("## 📊 計算結果（ボイル・シャルルの法則に基づく補正）")
            st.write(f"- 補正後終了圧力 P2_corr: **{P2_corr:.4f} MPa**")
            st.write(f"- 圧力変化量 ΔP: **{ΔP:.4f} MPa**")
            st.write(f"- 判定範囲: ±**{判定範囲:.4f} MPa**")
            st.markdown(f"### <span style='color:{色};'>判定結果: {合否}</span>", unsafe_allow_html=True)

            wb = openpyxl.load_workbook(TEMPLATE)
            ws = wb["気密試験記録"]

            def write(ws, cell, value):
                try:
                    ws[cell].value = value
                except AttributeError:
                    r = ws[cell].row
                    c = ws[cell].column
                    ws.cell(row=r, column=c, value=value)

            # Excel書き込み
            write(ws, "D3", 系統名)
            write(ws, "D4", 試験圧力)
            write(ws, "M4", 試験範囲)
            write(ws, "D5", 試験媒体)
            write(ws, "M5", 放置時間)
            write(ws, "D6", 使用機器No)
            write(ws, "M6", 測定場所)
            write(ws, "D8", 開始日時.strftime("%Y/%m/%d %H:%M"))
            write(ws, "M8", 終了日時.strftime("%Y/%m/%d %H:%M"))
            write(ws, "A10", f"{P1:.4f}")
            write(ws, "C10", f"{T1:.1f}")
            write(ws, "E10", f"{P2p:.4f}")
            write(ws, "G10", f"{T2:.1f}")
            write(ws, "J10", f"{P2_corr:.4f}")
            write(ws, "M10", f"{ΔP:.4f}")
            write(ws, "O10", f"±{判定範囲:.4f}")
            write(ws, "M11", 合否)
            write(ws, "E11", 試験実施者)

            output = BytesIO()
            wb.save(output)
            excel_data = output.getvalue()
            filename = f"気密試験記録_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

            b64 = base64.b64encode(excel_data).decode()
            href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{filename}">📥 Excelをダウンロード</a>'
            st.markdown(href, unsafe_allow_html=True)

            # --- Outlookメール送信 ---
            if st.session_state["email"] and st.session_state["app_password"]:
                try:
                    st.info("📤 Outlookメールに送信中...")
                    msg = MIMEMultipart()
                    msg["From"] = st.session_state["email"]
                    msg["To"] = st.session_state["email"]
                    msg["Subject"] = f"気密試験記録（{系統名 or '無題'}）"

                    body = MIMEText(f"判定結果: {合否}\n補正後圧力: {P2_corr:.4f} MPa\nΔP: {ΔP:.4f} MPa")
                    msg.attach(body)
                    attachment = MIMEApplication(excel_data)
                    attachment.add_header("Content-Disposition", "attachment", filename=filename)
                    msg.attach(attachment)

                    with smtplib.SMTP("smtp.office365.com", 587) as server:
                        server.starttls()
                        server.login(st.session_state["email"], st.session_state["app_password"])
                        server.send_message(msg)

                    st.success("✅ Outlookメールを送信しました。")

                except Exception as e:
                    st.error(f"📩 メール送信エラー: {e}")
            else:
                st.warning("⚠ メール設定が未登録です。上部で設定してください。")

        except Exception as e:
            st.error(f"⚠ エラーが発生しました: {e}")
